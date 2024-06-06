## Imports de base 
import pandas as pd 
from datetime import datetime
import recuperation_de_donnees
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, Border, Side
from openpyxl.formatting.rule import ColorScaleRule


#definition pour associe une colonnes du tableau (en lettre) a un chiffre ( qui est pour ce code une temperature)
#elle renvoie les lettres associées au chiffre 
def nombre_vers_lettre(n):
    """
    Cette fonction prend en entrée un nombre entier représentant une température et renvoie la lettre associée à ce nombre
    en fonction d'une convention. La lettre associée est utilisée pour identifier une colonne dans un tableau.
    La convention attribue des lettres aux nombres de manière à ce que les premiers nombres soient représentés par des
    lettres de l'alphabet, et lorsque les nombres dépassent 8, la lettre 'A' est préfixée, et lorsque les nombres dépassent
    34, la lettre 'B' est préfixée.
    arguments : 
        n (int): Un nombre entier représentant une température.
    returns:
        str: Une lettre associée au nombre donné.
    """
    n = int(n)
    if n < 9:
        lettres = chr(n + 82)  # Utiliser le code Unicode de 'A' (65) pour la logique de la première condition
    elif n > 8 and n < 35:
        lettres = 'A' + chr(n - 8 + 64)  # Utiliser 'A' en tant que première lettre et ajuster pour les autres lettres
    else:
        lettres = 'B' + chr(n - 34 + 64)  # Utiliser 'B' en tant que première lettre et ajuster pour les autres lettres
    return lettres



def associer_humidite_chiffre(humidite):
    """
    Cette fonction associe un pourcentage d'humidité à un chiffre spécifique. Elle prend en entrée un pourcentage
    d'humidité et renvoie le chiffre associé à ce pourcentage, en fonction d'une correspondance prédéfinie.
    arguments
        humidite (int): Le pourcentage d'humidité.
    returns:
        int: Le chiffre associé au pourcentage d'humidité, ou None si le pourcentage donné n'est pas trouvé dans la correspondance prédéfinie.
    """

    pourcentages = [0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54, 56, 58, 60, 62, 64, 66, 68, 70, 72, 74, 76, 78, 80, 82, 84, 86, 88, 90, 92, 94, 96, 98, 100]
    lignes_tab = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54]

    # Création du dictionnaire associant les éléments de liste1 à liste2
    association = dict(zip(pourcentages, lignes_tab))

    # Vérification si le chiffre est présent dans le dictionnaire et affichage du deuxième chiffre associé
    if humidite in association:
        deuxieme_chiffre = association[humidite]
        return deuxieme_chiffre
    else:
        print("Le chiffre n'est pas présent dans le dictionnaire.")


def fichier_excel(donnees, annee_cible):
    ##EXCLURE LES DONNEES DES ANNEES PAS CHOISIES PAR L'UTILISATEUR

    donnees_annee_cible = donnees[donnees['date'].dt.year == annee_cible] 

    ## SEPARER LES DONNES PAR MOIS SUR 1 ANNEE (l'année choisie)

    # Liste des noms des mois
    Mois = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    # Créer un dictionnaire pour stocker les DataFrames par mois
    donnees_par_mois = {}

    # Boucle pour itérer sur chaque mois de l'année
    for index, nom_mois in enumerate(Mois, start=1):
        #prendre les données pour le mois en cours
        donnees_mois = donnees_annee_cible[donnees_annee_cible['date'].dt.month == index]
        
        #stocker les données du mois dans le dictionnaire
        donnees_par_mois[nom_mois] = donnees_mois
        
        # Compter le nombre de données pour chaque température et l'humidité en fonction de la température
        donnees_par_mois[nom_mois] = donnees_par_mois[nom_mois].groupby(['temp_celsius', 'humidite']).size().reset_index(name='nbre_heures')

        # *3 pour avoir le nombre d'heures car on a des données que toutes les 3h
        donnees_par_mois[nom_mois]['nbre_heures'] = (donnees_par_mois[nom_mois]['nbre_heures'])*3


    #CREATION DU FICHIER EXCEL CONTENANT LES NOMBRES D'HEURES A UNE HUMIDITE DONNEE EN FONCTION DE LA TEMPERATURE

    #creation du fichier excel
    fichier_excel = Workbook()

    #creation des feuilles coorespondantes a chaque mois de l'année
    for m in Mois:
        fichier_excel.create_sheet(title=m)

    #creer une liste qui va contenir les pourcentage d'humidite par pas de 2
    pourcentages = []

    #creer une liste qui va contenir les lignes du tableau ou sont notées les valeurs de l'humidite 
    lignes_tab = []

    #supprimer la feuille par défaut
    fichier_excel.remove(fichier_excel["Sheet"])  

    #creation du tableau temperature/humidite dans chaque feuilles
    for Mois in fichier_excel.sheetnames:
        feuille = fichier_excel[Mois]
        border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        border_col = Border(top=Side(style=None), bottom = Side(style=None), left=Side(style='thin'), right=Side(style='thin'))
        feuille.cell(row=2, column=3, value="Température")
        feuille.cell(row=2, column=2, value="Humidité")
        feuille.merge_cells(start_row=2, start_column=3, end_row=2, end_column=63)

        #Ajouter la latitude et la longitude directement sur le fichier excel
        feuille.cell(row=1, column=1, value="Latitude")
        feuille.cell(row=1, column=2, value=donnees['latitude'][0])
        feuille.cell(row=1, column=4, value="Longitude")
        feuille.cell(row=1, column=5, value=donnees['longitude'][0]) 

        
        #mettre une legendes temperature pour les donnees a l'horizontale
        for col in range(3, 64):
            feuille.cell(row=2, column=col).border = border
            feuille.cell(row=3, column=col).border = Border(bottom=Side(style='thin'))

        #ecrire les en-têtes de colonnes pour les températures (pas de 1)
        for temp in range(-15, 46):
            feuille.cell(row=3, column=temp + 18, value=f"{temp}°C")

        #mettre un legendes pour la colonne d'humidite (a la verticale)
        for lignes in range(3,55):
            feuille.cell(row=lignes, column=2).border = border_col
            feuille.cell(row=2,column=2).border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            feuille.cell(row=54,column=2).border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        #ecrire les en-têtes de lignes pour les taux d'humidité (pas de 2)
        for humidite in range(0, 101, 2):
            feuille.cell(row=humidite // 2 + 4, column=2, value=f"{humidite}%")

        #mettre toutes les cases du tableau a zero 
        for lignes in range(4,55):
            for colonnes in range(3,64):
                feuille.cell(row=lignes,column=colonnes, value=0)

        for i in range(len(donnees_par_mois[Mois])):
            colonne_donnee = nombre_vers_lettre(donnees_par_mois[Mois]['temp_celsius'][i])
            ligne_donnee = associer_humidite_chiffre(donnees_par_mois[Mois]['humidite'][i])
            nom_case = colonne_donnee + str(ligne_donnee)
            feuille[nom_case] = donnees_par_mois[Mois]['nbre_heures'][i]

        color_scale_rule = ColorScaleRule(start_type="min", start_color=Color(rgb="FFFFFF"),end_type="max", end_color=Color(rgb="CD0000"))
        feuille.conditional_formatting.add('C4:BK54', color_scale_rule)

    #enregister le fichier excel
    fichier_excel.save(f"T_H_{donnees['nom_commune'][0]}_{annee_cible}.xlsx")
    return f"\"T_H_{donnees['nom_commune'][0]}_{annee_cible}.xlsx\""


def fichier_excel_moyenne(base_de_donnees):

    # Trouver les index des lignes correspondant au 29 février
    index_29_fevrier = base_de_donnees[(base_de_donnees['date'].dt.month == 2) & (base_de_donnees['date'].dt.day == 29)].index
    # Supprimer les lignes correspondant au 29 février
    base_de_donnees = base_de_donnees.drop(index_29_fevrier)

    annee_debut = base_de_donnees['date'].dt.year.iloc[0]
    annee_fin = base_de_donnees['date'].dt.year.iloc[-1]

    #Convertir la colonne 'date' en datetime
    base_de_donnees['date'] = pd.to_datetime(base_de_donnees['date'])
    # Créer une nouvelle colonne avec le mois et le jour uniquement
    base_de_donnees['mois_jour'] = base_de_donnees['date'].dt.strftime('%m-%d %H:%M:%S')
    
    # Convertir la colonne 'mois_jour' en datetime
    base_de_donnees['mois_jour'] = pd.to_datetime(base_de_donnees['mois_jour'], format='%m-%d %H:%M:%S')


    # Trier le DataFrame en fonction de la colonne 'mois_jour'
    base_de_donnees_sorted = base_de_donnees.sort_values(by='mois_jour')
    base_de_donnees_sorted['mois_jour'] = pd.to_datetime(base_de_donnees_sorted['mois_jour'], format='%m-%d %H:%M:%S')

    # Réinitialiser les index à zéro
    base_de_donnees_sorted = base_de_donnees_sorted.reset_index(drop=True)

    #Liste contenant tout les mois de l'année 
    Mois = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    # Créer un dictionnaire pour stocker les DataFrames par mois
    donnees_par_mois = {}

    # Boucle pour itérer sur chaque mois de l'année
    for index, nom_mois in enumerate(Mois, start=1):
        #prendre les données pour le mois en cours
        donnees_mois = base_de_donnees[base_de_donnees['mois_jour'].dt.month == index]
        #stocker les données du mois dans le dictionnaire
        donnees_par_mois[nom_mois] = donnees_mois
        # Compter le nombre de données pour chaque température et l'humidité en fonction de la température
        donnees_par_mois[nom_mois] = donnees_par_mois[nom_mois].groupby(['temp_celsius', 'humidite']).size().reset_index(name='nbre_heures')
        # *3 pour avoir le nombre d'heures car on a des données que toutes les 3h puis diviser par annee_finn-annee_debut afin de faire une "moyenne"
        donnees_par_mois[nom_mois]['nbre_heures'] = ((donnees_par_mois[nom_mois]['nbre_heures'])*3) / (annee_fin-annee_debut)

    #CREATION DU FICHIER EXCEL CONTENANT LES NOMBRES D'HEURES A UNE HUMIDITE DONNEE EN FONCTION DE LA TEMPERATURE
    #creation du fichier excel
    fichier_excel_moyenne = Workbook()

    #creation des feuilles coorespondantes a chaque mois de l'année
    for m in Mois:
        fichier_excel_moyenne.create_sheet(title=m)

    #supprimer la feuille par défaut
    fichier_excel_moyenne.remove(fichier_excel_moyenne["Sheet"])  

    #creation du tableau temperature/humidite dans chaque feuilles
    for Mois in fichier_excel_moyenne.sheetnames:
        feuille = fichier_excel_moyenne[Mois]
        border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        border_col = Border(top=Side(style=None), bottom = Side(style=None), left=Side(style='thin'), right=Side(style='thin'))
        feuille.cell(row=2, column=3, value="Température")
        feuille.cell(row=2, column=2, value="Humidité")
        feuille.merge_cells(start_row=2, start_column=3, end_row=2, end_column=63)

        #Ajouter la latitude et la longitude directement sur le fichier excel
        feuille.cell(row=1, column=1, value="Latitude")
        feuille.cell(row=1, column=2, value=base_de_donnees['latitude'][0])
        feuille.cell(row=1, column=4, value="Longitude")
        feuille.cell(row=1, column=5, value=base_de_donnees['longitude'][0]) 

        #mettre une legendes temperature pour les donnees a l'horizontale
        for col in range(3, 64):
            feuille.cell(row=2, column=col).border = border
            feuille.cell(row=3, column=col).border = Border(bottom=Side(style='thin'))

        #ecrire les en-têtes de colonnes pour les températures (pas de 1)
        for temp in range(-15, 46):
            feuille.cell(row=3, column=temp + 18, value=f"{temp}°C")

        #mettre un legendes pour la colonne d'humidite (a la verticale)
        for lignes in range(3,55):
            feuille.cell(row=lignes, column=2).border = border_col
            feuille.cell(row=2,column=2).border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            feuille.cell(row=54,column=2).border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        #ecrire les en-têtes de lignes pour les taux d'humidité (pas de 2)
        for humidite in range(0, 101, 2):
            feuille.cell(row=humidite // 2 + 4, column=2, value=f"{humidite}%")

        #mettre toutes les cases du tableau a zero 
        for lignes in range(4,55):
            for colonnes in range(3,64):
                feuille.cell(row=lignes,column=colonnes, value=0)

        for i in range(len(donnees_par_mois[Mois])):
            colonne_donnee = nombre_vers_lettre(donnees_par_mois[Mois]['temp_celsius'][i])
            ligne_donnee = associer_humidite_chiffre(donnees_par_mois[Mois]['humidite'][i])
            nom_case = colonne_donnee + str(ligne_donnee)
            feuille[nom_case] = donnees_par_mois[Mois]['nbre_heures'][i]

        color_scale_rule = ColorScaleRule(start_type="min", start_color=Color(rgb="FFFFFF"),end_type="max", end_color=Color(rgb="CD0000"))
        feuille.conditional_formatting.add('C4:BK54', color_scale_rule)

    #enregister le fichier excel
    fichier_excel_moyenne.save(f"T_H_MOYENNE_{base_de_donnees['nom_commune'][0]}_{annee_debut}_{annee_fin-1}.xlsx")
    return f"\"T_H_MOYENNE_{base_de_donnees['nom_commune'][0]}__{annee_debut}_{annee_fin-1}.xlsx\""
